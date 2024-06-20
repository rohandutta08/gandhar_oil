import time
import os
import openpyxl
import re
from tqdm import tqdm
from tabulate import tabulate
from framework.utils.credentials_loader import CredentialsLoader
from framework.business.recon.recon_workflow_loader import ReconWorkflowLoader
from framework.business.recon.recon_workflow import ReconIntegrationClient
from framework.utils.s3_util import download_file_from_s3


class ReconManager:
    def __init__(self, current_path, elt_instance_path):
        self.current_path = current_path
        self.elt_instance_path = elt_instance_path
        self.credentials_loader = CredentialsLoader(f"{self.elt_instance_path}/config/credentials.json")
        api_credentials = self.credentials_loader.load()

    def trigger(self):
        pass


class ReconWorkflowManager:
    def __init__(self,file_path):
        self.workflow_loader = ReconWorkflowLoader(file_path)
        self.workflow_request = self.workflow_loader.load()
        self.recon_request = self.workflow_loader.get_recon_request(self.workflow_request)
        
    def trigger_and_poll_workflow(self):
        recon_integration_client = ReconIntegrationClient(self.workflow_request)
        response = recon_integration_client.trigger_recon_workflow(self.recon_request)
        result = response.get('result',{})
        async_task_id = result.get('asyncTaskId')
        return self.get_workflow_status(recon_integration_client,async_task_id)
    
    def get_workflow_status(self,client,task_id):
        counter = 0
        with tqdm(total=100,desc="Recon Progress") as pbar:
            while True:
                api_response = client.get_workflow_current_status(task_id)
                result = api_response['result']
                status = result['task']['status']
                complete_percent = result['completePercent']
                if status not in ['PROCESSING', 'CREATED']:
                    break
                time.sleep(5)
                if (counter > complete_percent and complete_percent == 0):
                    counter = self.progress_update(pbar,counter,1)
                update_count = complete_percent - counter    
                if update_count > 33 :
                    counter = self.progress_update(pbar,counter,update_count) 
                else:
                    counter = self.progress_update(pbar,counter,1)    
                       
            self.progress_remain(pbar,counter)    
        return api_response               
                       
        
    def progress_update(self,pbar,counter,count):
        if counter < 95:
            pbar.update(count)
            counter += count
        return counter
              
    
    def progress_remain(self,pbar,counter):
        remain_count = 100 - counter
        if remain_count > 0:
            pbar.update(remain_count)      
                
                
                
    def generate_mismatch_field_report(self,wb, new_wb):
        sheet_name = 'Document Details (Inv CDN)'
        sheet = wb[sheet_name]

        # Create a new sheet within the existing workbook
        sheet_output = new_wb.active
        sheet_output.title = "MismatchError"

        # Copy headers up to column BI
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                sheet_output[cell.coordinate].value = cell.value
                if cell.coordinate == 'CJ1':
                    break
        
        split_values_count = {}
        sorted_rows = []

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            match_status = row[10].value
            mismatch_fields = row[14].value
    
            if match_status == 'Mismatch':
                split_values = re.split(',', mismatch_fields)
                for split_value in split_values:
                    sorted_rows.append((split_value, row))
        
    
        sorted_rows.sort(key=lambda x: x[0])

        for sorted_row in sorted_rows:
            split_value, row = sorted_row
            split_values_count[split_value] = split_values_count.get(split_value, 0) + 1
            new_row = [cell.value for cell in row]
            new_row[14] = split_value  
            sheet_output.append(new_row)
        

        # Create a new sheet for error messages and row counts
        sheet_errors = new_wb.create_sheet(title="Mismatch Counts")

        # # Write headers for the new sheet
        sheet_errors["A1"] = "Mismatch Field"
        sheet_errors["B1"] = "Distinct Document Count"
        sheet_errors["C1"] = "Sheet Row No"

        for idx, (split_value, row_indices) in enumerate(split_values_count.items(), start=2):
            sheet_errors[f"A{idx}"] = split_value
            row_count_cell = f"B{idx}"
            matching_rows = [row[0].row for row in sheet_output.iter_rows() if row[14].value == split_value]
            if matching_rows:
                size = len(matching_rows)
                cell_range = f"A{matching_rows[0]}:CJ{matching_rows[size-1]}"
                sheet_errors[row_count_cell].value = f"{len(matching_rows)}"
                sheet_errors[row_count_cell].hyperlink = f"#MismatchError!{cell_range}"
            else:
                print('no matching rows')
                sheet_errors[row_count_cell].value = "Split value not found in MismatchError"

        new_wb.move_sheet(sheet_errors, offset=-1)           
         
    def get_tax_diff_key(self, tax_diff):
        tax_diff_ranges = {
            0: 0,
            1: 1,
            10: 2,
            100: 3,
            1000: 4,
            10000: 5,
            100000: 6
        }
    
        for diff_range, key in tax_diff_ranges.items():
            if tax_diff < diff_range:
                return key
        return 7 
    
         
    def generate_tax_diff(self,wb,new_wb):

        sheet_name = 'Document Details (Inv CDN)'
        sheet = wb[sheet_name]

        # Create a new sheet within the existing workbook
        sheet_output = new_wb.create_sheet(title="TaxDiffDetails")

        # Copy headers up to column BI
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                sheet_output[cell.coordinate].value = cell.value
                if cell.coordinate == 'CK1':
                    break
        
        # Create a dictionary to store split values and their row counts
        split_values_count = {}
        sorted_rows = []
        tax_diff_map = {
            0: '< 0 Rs Difference',
            1: '0 - 1 Rs Difference',
            2: '1-10 Rs Difference',
            3: '11- 100 Rs Difference',
            4: '101- 1000 Rs Difference',
            5: '1001 - 100000 Rs Difference',
            6: '100001 - 1000000 Rs Difference',
            7: '> 1000000'
        }
        # Iterate through each row starting from row 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Get the value in column B for the current row
            match_status = row[10].value
            tax_diff = row[31].value
    
            if match_status == 'Mismatch':
                sorted_rows.append((self.get_tax_diff_key(tax_diff), row))
        
    
        # Sort the rows based on value_B
        sorted_rows.sort(key=lambda x: x[0])


        for sorted_row in sorted_rows:
            split_value, row = sorted_row
            split_values_count[split_value] = split_values_count.get(split_value, 0) + 1
            new_row = [cell.value for cell in row]
            new_row[0] = tax_diff_map[split_value]  
            sheet_output.append(new_row)
        

        # Create a new sheet for error messages and row counts
        sheet_errors = new_wb.create_sheet(title="TaxDiff Counts")

        # # Write headers for the new sheet
        sheet_errors["A1"] = "Tax Diff range"
        sheet_errors["B1"] = "Distinct Document Count"

        # Write split values and their row counts to the new sheet
        for idx, (split_value, row_indices) in enumerate(split_values_count.items(), start=2):
            sheet_errors[f"A{idx}"] = tax_diff_map[split_value]
            # sheet_errors[f"C{idx}"] = row_indices
            # Create hyperlinks for row count values
            # link_cell = f"C{idx}"
            row_count_cell = f"B{idx}"
            # Link to the first row in MyCustomSheetName where the split value occurs if it exists
            matching_rows = [row[0].row for row in sheet_output.iter_rows() if row[0].value == tax_diff_map[split_value]]
            if matching_rows:
                size = len(matching_rows)
                cell_range = f"A{matching_rows[0]}:CJ{matching_rows[size-1]}"
                selected_cells = f"A{', A'.join(map(str, matching_rows))}"
        
                # sheet_errors[link_cell].value = selected_cells
                sheet_errors[row_count_cell].value = f"{len(matching_rows)}"
                sheet_errors[row_count_cell].hyperlink = f"#TaxDiffDetails!{cell_range}"
            else:
                print('no matching rows')
                sheet_errors[row_count_cell].value = "Split value not found in TaxDiffDetails"   
        
        new_wb.move_sheet(sheet_errors, offset=-1)         
                
    
     
                        


    
    def generate_tax_val_diff(self,wb, new_wb):

        sheet_name = 'Document Details (Inv CDN)'
        sheet = wb[sheet_name]

        # Create a new sheet within the existing workbook
        sheet_output = new_wb.create_sheet(title="TaxableValDiffDetails")

        # Copy headers up to column BI
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                sheet_output[cell.coordinate].value = cell.value
                if cell.coordinate == 'CK1':
                    break
        
        # Create a dictionary to store split values and their row counts
        split_values_count = {}
        sorted_rows = []
        tax_diff_map = {
            0: '< 0 Rs Difference',
            1: '0 - 1 Rs Difference',
            2: '1-10 Rs Difference',
            3: '11- 100 Rs Difference',
            4: '101- 1000 Rs Difference',
            5: '1001 - 100000 Rs Difference',
            6: '100001 - 1000000 Rs Difference',
            7: '> 1000000'
        }
        # Iterate through each row starting from row 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Get the value in column B for the current row
            match_status = row[10].value
        
            if match_status == 'Mismatch':
                govt_taxable_val = row[29].value
                pr_taxable_val = row[30].value
                tax_diff = govt_taxable_val - pr_taxable_val
                sorted_rows.append((self.get_tax_diff_key(tax_diff), row))
        
    
        # Sort the rows based on value_B
        sorted_rows.sort(key=lambda x: x[0])


        for sorted_row in sorted_rows:
            split_value, row = sorted_row
            split_values_count[split_value] = split_values_count.get(split_value, 0) + 1
            new_row = [cell.value for cell in row]
            new_row[0] = tax_diff_map[split_value]  
            sheet_output.append(new_row)
        

        # Create a new sheet for error messages and row counts
        sheet_errors = new_wb.create_sheet(title="TaxableValDiff Counts")

        # # Write headers for the new sheet
        sheet_errors["A1"] = "Tax Diff range"
        sheet_errors["B1"] = "Distinct Document Count"

        # Write split values and their row counts to the new sheet
        for idx, (split_value, row_indices) in enumerate(split_values_count.items(), start=2):
            sheet_errors[f"A{idx}"] = tax_diff_map[split_value]
            row_count_cell = f"B{idx}"
            matching_rows = [row[0].row for row in sheet_output.iter_rows() if row[0].value == tax_diff_map[split_value]]
            if matching_rows:
                size = len(matching_rows)
                cell_range = f"A{matching_rows[0]}:CJ{matching_rows[size-1]}"
                sheet_errors[row_count_cell].value = f"{len(matching_rows)}"
                sheet_errors[row_count_cell].hyperlink = f"#TaxableValDiffDetails!{cell_range}"
            else:
                print('no matching rows')
                sheet_errors[row_count_cell].value = "Split value not found in TaxableValDiffDetails"

        new_wb.move_sheet(sheet_errors, offset=-1)  
        
        
    def generate_cross_section_report(self,wb,new_wb):

        sheet_name = 'Document Details (Inv CDN)'
        sheet = wb[sheet_name]

        
        # Create a dictionary to store split values and their row counts
        split_values_count = {}
        sorted_rows = []
        cross_section_map = {
            'B2B':'3',
            'CDN':'4',
            'B2BA':'5',
            'CDNA':'6'
        }
        cell_address_map = {
            'B2B': 'C',
            'CDN': 'D',
            'B2BA': 'E',
            'CDNA': 'F'
        }
        # Iterate through each row starting from row 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Get the value in column B for the current row
            match_status = row[10].value
        
            if match_status == 'Mismatch':
                govt_section_name = row[44].value
                pr_section_name = row[45].value
                cell_addr = cell_address_map[govt_section_name]
                pr_index = cross_section_map[pr_section_name]
                section_key = cell_addr+pr_index
                sorted_rows.append(section_key)


        for section_key in sorted_rows:
            split_values_count[section_key] = split_values_count.get(section_key, 0) + 1
        

        # Create a new sheet for error messages and row counts
        sheet_errors = new_wb.create_sheet(title="Cross Section")

        # # Write headers for the new sheet
        sheet_errors.merge_cells('A3:A6')
        sheet_errors['A3'] = 'Clear SR/PR'
        sheet_errors.merge_cells('C1:F1')
        sheet_errors['C1'] = 'GSTIN'
        sheet_errors['B3'] = 'B2B'
        sheet_errors['B4'] = 'CDN'
        sheet_errors['B5'] = 'B2BA'
        sheet_errors['B6'] = 'CDNA'
        sheet_errors['C2'] = 'B2B'
        sheet_errors['D2'] = 'CDN'
        sheet_errors['E2'] = 'B2BA'
        sheet_errors['F2'] = 'CDNA'
    
        for key in split_values_count:
            sheet_errors[key] = split_values_count[key]  
        
    
    def generate_custom_report(self,local_file):
        with tqdm(total=100, desc="Recon Custom Report Progress") as pbar:
            wb = openpyxl.load_workbook(local_file)
            pbar.update(10)

            new_wb = openpyxl.Workbook()
            self.generate_mismatch_field_report(wb,new_wb)
            pbar.update(20)
            self.generate_tax_diff(wb,new_wb)
            pbar.update(20)
            self.generate_tax_val_diff(wb,new_wb)
            pbar.update(20)
            self.generate_cross_section_report(wb,new_wb)
            pbar.update(20)

            base_file_name = os.path.basename(local_file)
            folder_path = os.path.dirname(local_file)
            custom_name = f'custom_{base_file_name}'
            custom_path = os.path.join(folder_path, custom_name)
            new_wb.save(custom_path)
            pbar.update(10)
            wb.close()
        
        
    def generate_summary_report(self,local_file):
        wb = openpyxl.load_workbook(local_file)
        sheet_name = 'Overall Summary'
        sheet = wb[sheet_name]
        summary_data = []
        with tqdm(total=100, desc="Recon Summary Progress") as pbar:
            for num in range(13, 20):
                match_status_ref = f'A{num}'
                govt_doc_count_ref = f'E{num}'
                govt_taxable_ref = f'F{num}'
                pr_doc_count_ref = f'H{num}'
                pr_taxable_ref = f'I{num}'
                summary_data.append([f'{str(sheet[match_status_ref].value)}',f'{str(sheet[govt_doc_count_ref].value)}',f'{str(sheet[govt_taxable_ref].value)}',f'{str(sheet[pr_doc_count_ref].value)}',f'{str(sheet[pr_taxable_ref].value)}'])
                pbar.update(10)
            pbar.update(30)     
        return summary_data  
    
    def prepare_summary(self,local_file):
        summary_data = self.generate_summary_report(local_file)
        headers = ["Match Status","2b doc count", "2b taxableValue","pr doc count","pr taxableValue"]
        print(tabulate(summary_data, headers=headers, tablefmt="grid"))
        
    
    def process_summary_and_custom_report(self,api_response,report_folder_path):
        status = api_response['result']['task']['status']
        if status == 'SUCCESS':
            report_result = api_response['result']["task"]["response"]["reconReportResults"]  
            report_url = report_result[0]['presignedUrl']
            local_file = download_file_from_s3(report_url,report_folder_path)
            self.prepare_summary(local_file)
            self.generate_custom_report(local_file)
            
                                            