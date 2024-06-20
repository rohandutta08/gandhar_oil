from framework.business.ingestion.ingestion_client import IngestionClient
from framework.business.ingestion.ingestion_workflow_loader import IngestionWorkflowLoader
from framework.utils.user_service_client import UserServiceClient
import os
from framework.utils.s3_util import download_file_from_s3
import openpyxl
import re
import pandas as pd
from openpyxl.utils import get_column_letter
import time
from tabulate import tabulate
from tqdm import tqdm

class IngestionManager:
    def __init__(self, current_folder,file_path):
        file_path = f"{current_folder}{file_path}"
        metadata_path = f"{current_folder}/framework/business/ingestion/config/ingestion-config.json"
        self.workflow_loader = IngestionWorkflowLoader(file_path,metadata_path)
        self.workflow_request = self.workflow_loader.load()
        self.pr_ingest_file_path = f"{current_folder}{self.workflow_request.prIngestFilePath}"
        self.sales_ingest_file_path = f"{current_folder}{self.workflow_request.salesIngestFilePath}"
        self.pr_return_period_start = self.workflow_request.prReturnPeriodStart
        self.pr_return_period_end = self.workflow_request.prReturnPeriodEnd
        self.sales_return_period_start = self.workflow_request.salesReturnPeriodStart
        self.sales_return_period_end = self.workflow_request.salesReturnPeriodEnd

    def trigger(self, ingestion_folder_path):
        user_id = self.get_user_id()
        file_types = ["PURCHASE", "SALES"]
        processing_summary = {}
        summary_headers = ["GSTIN","Docs", "Rows","Errors","No Errors"]
        for file_type in file_types:
            progress_counter = 0
            with tqdm(total=100, desc=f"{file_type} Ingestion Progress") as pbar:
                progress_counter = self.progress_update(pbar,progress_counter,1)

                ingest_file_path, return_period_start, return_period_end = self.get_ingest_parameters(file_type)

                presigned_response = self.get_presigned_url(ingest_file_path, file_type,user_id)
                progress_counter = self.progress_update(pbar,progress_counter,10)

                ingestion_client = IngestionClient(self.workflow_request, user_id)
                ingestion_client.upload_to_s3(presigned_response['payload'], ingest_file_path)
                url = presigned_response["payload"].partition('?')[0]
                progress_counter = self.progress_update(pbar,progress_counter,10)

                activity_response = self.create_activity(ingestion_client, url, ingest_file_path, file_type, return_period_start, return_period_end)
                activity_id = activity_response['payload']
                progress_counter = self.progress_update(pbar,progress_counter,10)

                activity_status_response = self.poll_activity_status(ingestion_client, activity_id, pbar, progress_counter)
                activity_status = activity_status_response['payload']['status']
                progress_counter = self.progress_update(pbar,progress_counter,10)

                if activity_status == 'FAILURE':
                    self.handle_failure(activity_status_response,ingestion_folder_path)

                self.progress_remain(pbar,progress_counter)

                processing_summary[file_type] = self.generate_error_summary_report(activity_status_response['payload']['gstinprocessingStats'])

            print(f'Ingestion summary for {file_type}')
            print(tabulate(processing_summary[file_type], headers=summary_headers, tablefmt="grid"))
            
     
    def progress_update(self,pbar,counter,count):
        if counter < 95:
            pbar.update(count)
        counter += count
        return counter
            
        
    
    def progress_remain(self,pbar,counter):
        remain_count = 100 - counter
        if remain_count > 0:
            pbar.update(remain_count)
    
    
    def get_ingest_parameters(self, file_type):
        if file_type == 'SALES':
            return self.sales_ingest_file_path, self.sales_return_period_start, self.sales_return_period_end
        return self.pr_ingest_file_path, self.pr_return_period_start, self.pr_return_period_end
    
    def get_user_id(self):
        user_service_client = UserServiceClient()
        return user_service_client.get_user_id(self.workflow_request.authToken)
    
    def get_presigned_url(self, ingest_file_path, file_type,user_id):
        ingestion_client = IngestionClient(self.workflow_request, user_id)
        return ingestion_client.get_presigned_url(ingest_file_path, file_type)
    
    def create_activity(self, ingestion_client, url, ingest_file_path, file_type, return_period_start, return_period_end):
        file_name = os.path.basename(ingest_file_path)
        return ingestion_client.create_activity(url, file_name, file_type, return_period_start, return_period_end)
    
    def poll_activity_status(self, ingestion_client, activity_id, pbar, progress_counter):
        while True:
            time.sleep(1)
            progress_counter  = self.progress_update(pbar,progress_counter,2)

            activity_status_response = ingestion_client.get_activity_status(activity_id)
            activity_status = activity_status_response['payload']['status']
            activity_state = activity_status_response['payload']['state']

            if activity_status in ["FAILURE", "SUCCESS"] and activity_state in ["ACTIVITY_COMPLETED", "SYSTEM_FEEDBACK", "UNPLANNED_ERROR"]:
                return activity_status_response 
            
    
    def handle_failure(self, activity_status_response, ingestion_folder_path):
        payload = activity_status_response['payload']
        error_sheets = payload['errorFileBySheets']
        if len(error_sheets) >= 1:
            error_sheet = error_sheets[0]
            local_file_path = download_file_from_s3(error_sheet["fileUrl"], ingestion_folder_path)
            base_name, ext = os.path.splitext(local_file_path)
            file_name = f"{base_name}.xlsx"
            output_file_path = os.path.join(ingestion_folder_path, file_name)
            df = pd.read_csv(local_file_path)
            df.to_excel(output_file_path, index=False)
            self.generate_custom_report(output_file_path)                
         
        
    
    def generate_custom_report(self, file_path):
        # Load the Excel workbook
        wb = openpyxl.load_workbook(file_path) 
        
        # Get the active sheet
        sheet = wb.active
        # Create a new sheet within the existing workbook
        sheet_output = wb.create_sheet(title="PurchaseError")

        # Copy headers up to column BI
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                sheet_output[cell.coordinate].value = cell.value
                if cell.coordinate == 'BI1':
                    break
        
        # Create a dictionary to store split values and their row counts
        split_values_count = {}
        sorted_rows = []

        # Iterate through each row starting from row 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Get the value in column B for the current row
            value_B = row[1].value
    
            # Split the value and create rows accordingly
            split_values = re.split(r', (?=\[)', value_B)
            for split_value in split_values:
                # Append the row data to the list for sorting
                sorted_rows.append((split_value, row))

        # Sort the rows based on value_B
        sorted_rows.sort(key=lambda x: x[0])

        # Append sorted rows to the output sheet
        for sorted_row in sorted_rows:
            split_value, row = sorted_row
            split_values_count[split_value] = split_values_count.get(split_value, 0) + 1
            new_row = [cell.value for cell in row]
            new_row[1] = split_value  # Update the value in column B
            sheet_output.append(new_row)
        

        # Create a new sheet for error messages and row counts
        sheet_errors = wb.create_sheet(title="Error Counts")

        # Write headers for the new sheet
        sheet_errors["A1"] = "Error Message"
        sheet_errors["B1"] = "Distinct Document Count"
        # sheet_errors["C1"] = "Sheet Row No"

        # Write split values and their row counts to the new sheet
        for idx, (split_value, row_indices) in enumerate(split_values_count.items(), start=2):
            sheet_errors[f"A{idx}"] = split_value
            # sheet_errors[f"C{idx}"] = row_indices
            # Create hyperlinks for row count values
            # link_cell = f"C{idx}"
            row_count_cell = f"B{idx}"
            # Link to the first row in MyCustomSheetName where the split value occurs if it exists
            matching_rows = [row[0].row for row in sheet_output.iter_rows() if row[1].value == split_value]
            if matching_rows:
                size = len(matching_rows)
                cell_range = f"A{matching_rows[0]}:BI{matching_rows[size-1]}"
                selected_cells = f"A{', A'.join(map(str, matching_rows))}"
        
                # sheet_errors[link_cell].value = selected_cells
                sheet_errors[row_count_cell].value = f"{len(matching_rows)}"
                sheet_errors[row_count_cell].hyperlink = f"#PurchaseError!{cell_range}"
            else:
                sheet_errors[row_count_cell].value = "Split value not found in PurchaseError"

        wb.move_sheet(sheet_errors, offset=-1)
        # Save the updated workbook
        wb.save(file_path)
        # Close the workbook
        wb.close()
        
        
    def generate_error_summary_report(self,processing_stats):
        summary_data = []
        for processing_stat in processing_stats:
            gstin = processing_stat['gstin']
            docs_count = processing_stat['totalDocumentCount']
            processed_docs_count = processing_stat['processedRecords']
            error_count = processing_stat['invalidRecords']
            valid_docs_count = processing_stat['validRecords']
            summary_data.append([f'{str(gstin)}',f'{str(docs_count)}',f'{str(processed_docs_count)}',f'{str(error_count)}',f'{str(valid_docs_count)}'])     
        return summary_data
            
        






       
